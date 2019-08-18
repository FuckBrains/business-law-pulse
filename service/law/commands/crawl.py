# -*- coding: utf-8 -*-

from core.utils import REDIS
from core.utils.command import BaseCommand

from core.utils.setup import setup_redis,setup_mongo
setup_redis()
setup_mongo()

import json,datetime,time
import re,requests

from bs4 import BeautifulSoup
from urllib.parse import urljoin
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill,Font,colors

from service.law.models import LawDeal,LawClient,LawFirm
from service.law.models import EmbeddedDealClient,EmbeddedDealFirm



def get_response(url, headers={}):
    headers.update({
        'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36',
    })

    tries = 10
    response = None

    while tries:
        try:
            response = requests.get(url,headers=headers,timeout=10)
            break
        except Exception as exc:
            print('[Failure] %s' % url, flush=True)
            tries = tries - 1
            time.sleep(3)

    if not response:
        raise Exception('Over Tries')

    return response



def crawl_boom():
    records = {}

    entry = {
        'en': 'https://www.boom.com/jsp/en/market_info/ipo/listed.jsp',
        'cn': 'https://www.boom.com/jsp/gb/market_info/ipo/listed.jsp',
    }

    def crawl_lang(lang):
        response = get_response(entry[lang])
        soup = BeautifulSoup(response.text,'lxml')
        trs = soup.find(id='ipoListContainer').find('table').find_all('tr')

        for tr in trs[1:-2]:
            tds = tr.find_all('td')
            code = tds[1].text.strip()

            if not records.get(code,{}):
                records[code] = { 'date': '', 'area': '', 'website': '', 'en': {}, 'cn': {} }

            records[code]['date'] = datetime.datetime.strptime(tds[0].text,'%Y/%m/%d' if lang=='en' else '%Y年 %m月 %d日').strftime('%Y-%m-%d')

            records[code][lang] = {
                'name': tds[2].text.strip(),
                'industries': [item.strip() for item in tds[3].text.split(',')],
                'url': urljoin(entry[lang],tds[2].find('a').attrs['href']),
                'description': '',
                'prospectus': '',
            }

    crawl_lang('en')
    crawl_lang('cn')

    codes = list(records.keys())
    codes.sort()
    index = 0
    for code in codes:
        index = index + 1
        print('%s / %s => %s%%' % (index,len(codes),index*100//len(codes)), end='\r', flush=True)

        if REDIS['app'].exists('boom:%s' % code):
            continue

        record = records[code]
        for lang in ('en','cn'):
            response = get_response(record[lang]['url'])
            soup = BeautifulSoup(response.text,'lxml')
            trs = soup.find(class_='IPOInfoTable').find_all('table')[-1].find_all('tr')
            record[lang]['prospectus'] = trs[-1].find_all('td')[0].find('a').attrs['href']
            if len(trs) > 3:
                record[lang]['description'] = trs[2].text.strip().replace('\r','').replace('\n','')

            if lang == 'en':
                record['website'] = trs[-1].find_all('td')[-1].find('a').attrs['href']

                if record['en']['description'].find('PRC') != -1:
                    record['area'] = 'PRC'
                elif record['en']['description'].find('Hong Kong') != -1:
                    record['area'] = 'HK'

        # value
        en_value_url = 'http://www.etnet.com.hk/www/eng/stocks/ci_ipo_detail.php?code=%s' % code
        cn_value_url = 'http://www.etnet.com.hk/www/sc/stocks/ci_ipo_detail.php?code=%s' % code
        record['en']['value_url'] = en_value_url
        record['cn']['value_url'] = cn_value_url
        record['value'] = ''
        response = get_response(en_value_url)
        soup = BeautifulSoup(response.text,'lxml')
        div = soup.find('div',text='USE OF PROCEEDS')
        if div:
            tr = div.find_next_sibling().find('tr')
            if tr.text.find('would be') != -1:
                pattern = re.search('would be(?P<value>.+)(,|\.)\s', tr.text).groupdict()
                record['value'] = pattern['value'].strip()
                print(tr.text, flush=True)
                print(record['value'], flush=True)

        # solicitors
        en_solicitor_url = 'http://mpf.etnet.com.hk/www/eng/stocks/realtime/quote_ci_brief.php?code=%s' % code
        record['en']['solicitor_url'] = en_solicitor_url
        record['en']['solicitors'] = []
        response = get_response(en_solicitor_url)
        soup = BeautifulSoup(response.text,'lxml')
        node = soup.find('td',text='Solicitors')
        if node:
            tds = node.find_next_sibling().find_all('td')
            for td in tds:
                name = td.text[4:].strip()
                if name:
                    record['en']['solicitors'].append(name)

        cn_solicitor_url = 'http://mpf.etnet.com.hk/www/sc/stocks/realtime/quote_ci_brief.php?code=%s' % code
        record['cn']['solicitor_url'] = cn_solicitor_url
        record['cn']['solicitors'] = []
        response = get_response(cn_solicitor_url)
        response.encoding = 'utf8'
        soup = BeautifulSoup(response.text,'lxml')
        node = soup.find('td',text='律师')
        if node:
            tds = node.find_next_sibling().find_all('td')
            for td in tds:
                name = td.text.strip()
                if name:
                    record['cn']['solicitors'].append(name)

        REDIS['app'].set('boom:%s' % code, json.dumps(record), ex=60*60*24*30)



def crawl_cfi():
    records = {}
    for i in range(1,33):
        entry = 'http://quote.cfi.cn/cache_image/node224_%s___________0_0_%s_false.js' % (i,i)
        response = get_response(entry)
        content = response.text
        content = content[content.find('innerHTML')+11:-2]
        soup = BeautifulSoup(content,'lxml')
        trs = soup.find_all('tr')[1:-2]
        for tr in trs:
            tds = tr.find_all('td')
            code = tds[0].find('a').attrs['href'].split('_')[-1].split('.')[0]
            records[code] = {
                'name': tds[0].text.strip(),
                'date': datetime.datetime.strptime(tds[1].text,'%Y%m%d').strftime('%Y-%m-%d'),
            }

    codes = list(records.keys())
    codes.sort()
    index = 0
    for code in codes:
        index = index + 1
        print('%s / %s => %s%%' % (index,len(codes),index*100//len(codes)), end='\r', flush=True)

        if REDIS['app'].exists('cfi:%s' % code):
            continue

        record = records[code]
        response = get_response('http://quotes.money.163.com/f10/gszl_%s.html' % code)
        soup = BeautifulSoup(response.text,'lxml')
        trs = soup.find(class_='col_l_01').find('table').find_all('tr')
        record.update({
            'name_en': trs[3].find_all('td')[1].text.strip(),
            'name_cn': trs[2].find_all('td')[1].text.strip(),
            'phone': trs[2].find_all('td')[3].text.strip(),
            'email': trs[3].find_all('td')[3].text.strip(),
            'address': trs[1].find_all('td')[3].text.strip(),
            'website': trs[8].find_all('td')[1].text.strip(),
            'prospectus': [],
        })

        description = ''
        for tr in trs[-2:]:
            if tr.find_all('td')[0].text in ('经营范围','公司沿革'):
                if 'title' in tr.find_all('td')[1].attrs:
                    description += tr.find_all('td')[1].attrs['title']
                else:
                    description += tr.find_all('td')[1].text.strip()

        record.update({ 'description': description.replace('\r','').replace('\n','') })

        trs = soup.find(class_='col_r_01').find('table').find_all('tr')
        record.update({'value': trs[6].find_all('td')[1].text.strip() })

        prospectus = record['prospectus']
        sina = 'http://vip.stock.finance.sina.com.cn/corp/go.php/vISSUE_RaiseExplanation/stockid/%s.phtml' % code
        response = get_response(sina)
        response.encoding = 'gb2312'
        soup = BeautifulSoup(response.text,'lxml')
        if soup.find(class_='datelist'):
            for link in soup.find(class_='datelist').find_all('a'):
                prospectus.append({
                    'title': link.text,
                    'url': urljoin(sina,link.attrs['href']),
                })

            years = [item for item in re.findall('\d\d\d\d-\d\d-\d\d',soup.find(class_='datelist').text)]
            for i in range(0,len(years)):
                prospectus[i]['year'] = years[i]

        firm_url = 'http://app.finance.ifeng.com/data/stock/gsjj.php?symbol=%s' % code
        record['firm_url'] = firm_url
        response = get_response(firm_url)
        response.encoding = 'utf8'
        soup = BeautifulSoup(response.text,'lxml')
        td = soup.find('td',text='律师事务所').find_next_sibling()
        record['firm'] = td.text.strip()

        REDIS['app'].set('cfi:%s' % code, json.dumps(record), ex=60*60*24*30)



def crawl_cbonds():
    entry = {
        2017: 'http://em.cbonds.com/rankings/item/431',
        2016: 'http://cbonds.com/rankings/item/375',
        2015: 'http://cbonds.com/rankings/item/333',
    }

    records = {}
    for year in entry.keys():
        response = get_response(entry[year])
        soup = BeautifulSoup(response.text,'lxml')
        for li in soup.find(class_='js_show_prnt').find('ul').find_all('li'):
            if not li.find('div'):
                continue

            firm = li.find(class_='cb_left txt').text

            ul = li.find(class_='emtnt_list')
            for _li in ul.find_all('li'):
                title = _li.text.strip()
                url = _li.find('a').attrs['href']
                id = url.split('/')[-1]
                records[id] = { 'title': title, 'url': url, 'year': year }

    ids = list(records.keys())
    ids.sort()
    index = 0
    for id in ids:
        index = index + 1
        print('%s / %s => %s%%' % (index,len(ids),index*100//len(ids)), end='\r', flush=True)

        if REDIS['app'].exists('cbonds:%s' % id):
            continue

        record = records[id]
        record.update({
            'country': '',
            'amount': '',
            'borrower': '',
            'bookrunner': '',
            'advisor': {
                'issuer': { 'international': '', 'domestic': '', 'listing': '' },
                'arranger': { 'international': '', 'domestic': '', 'listing': '' },
            },
        })

        response = get_response(record['url'])
        soup = BeautifulSoup(response.text,'lxml')
        trs = soup.find_all('tr')
        for tr in trs:
            if len(tr.find_all('th')) == 5 and tr.find_all('th')[1].text.startswith('Country of risk'):
                tds = tr.parent.parent.find_all('tr')[1].find_all('td')
                record['country'] = tds[1].text
                record['amount'] = tds[3].text
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Borrower'):
                record['borrower'] = tr.find_all('td')[1].text
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Bookrunner'):
                record['bookrunner'] = [item.strip() for item in tr.find_all('td')[1].text.split(',')]
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Issuer Legal Adviser (International law)'):
                record['advisor']['issuer']['international'] = tr.find_all('td')[1].text
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Issuer Legal Adviser (Domestic law)'):
                record['advisor']['issuer']['domestic'] = tr.find_all('td')[1].text
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Issuer Legal Adviser (Listing law)'):
                record['advisor']['issuer']['listing'] = tr.find_all('td')[1].text
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Arranger Legal Adviser (International law)'):
                record['advisor']['arranger']['international'] = tr.find_all('td')[1].text
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Arranger Legal Adviser (Domestic law)'):
                record['advisor']['arranger']['domestic'] = tr.find_all('td')[1].text
            elif len(tr.find_all('td')) == 2 and tr.find_all('td')[0].text.startswith('Arranger Legal Adviser (Listing law)'):
                record['advisor']['arranger']['listing'] = tr.find_all('td')[1].text

        REDIS['app'].set('cbonds:%s' % id, json.dumps(record), ex=60*60*24*30)



def export_excel():
    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'Code'
    ws['B1'] = 'Date'
    ws['C1'] = 'Website'
    ws['D1'] = 'Area'
    ws['E1'] = 'Name EN'
    ws['F1'] = 'Industries EN'
    ws['G1'] = 'Description EN'
    ws['H1'] = 'Prospectus EN'
    ws['I1'] = 'URL EN'
    ws['J1'] = 'Name CN'
    ws['K1'] = 'Industries CN'
    ws['L1'] = 'Description CN'
    ws['M1'] = 'Prospectus CN'
    ws['N1'] = 'URL CN'

    index = 1
    for key in REDIS['app'].scan_iter('boom:*'):
        record = json.loads(REDIS['app'].get(key))
        code = key.split(':')[-1]

        index = index + 1

        ws['A%s' % index] = code
        ws['B%s' % index] = record['date']
        ws['C%s' % index] = record['website']
        ws['D%s' % index] = record['area']
        ws['E%s' % index] = record['en']['name']
        ws['F%s' % index] = '\n'.join(record['en']['industries'])
        ws['G%s' % index] = record['en']['description']
        ws['H%s' % index] = record['en']['prospectus']
        ws['I%s' % index] = record['en']['url']
        ws['J%s' % index] = record['cn']['name']
        ws['K%s' % index] = '\n'.join(record['cn']['industries'])
        ws['L%s' % index] = record['cn']['description']
        ws['M%s' % index] = record['cn']['prospectus']
        ws['N%s' % index] = record['cn']['url']

    wb.save('boom.xlsx')



"""
def batch_boom_input():
    ##########################################################################################
    # load constants
    ##########################################################################################
    constants = { 'categories': [], 'industries': [], 'areas': [], 'parties': [] }
    constants['categories'] = json.loads(REDIS['app'].get('law:config:categories'))
    constants['industries'] = json.loads(REDIS['app'].get('law:config:industries'))
    constants['areas'] = json.loads(REDIS['app'].get('law:config:areas'))
    constants['parties'] = json.loads(REDIS['app'].get('law:config:parties'))

    ##########################################################################################
    # validate excel records
    ##########################################################################################
    wb = load_workbook(filename='boom.xlsx')
    ws = wb.active
    keys = [
        'code','date','website','area','value',
        'name_en','industries_en','description_en','prospectus_en','url_en',
        'name_cn','industries_cn','description_cn','prospectus_cn','url_cn',
        'solicitor_en','solicitor_en_url',
        'solicitor_cn','solicitor_cn_url',
    ]

    error = False
    records = []
    for r in range(2,ws.max_row+1):
        values = []
        for c in range(1,ws.max_column+1):
            cell = ws.cell(row=r,column=c)
            values.append(cell.value)

        record = dict(zip(keys,values))
        record['industries_en'] = [item for item in record['industries_en'].split('\n')]
        record['industries_cn'] = [item for item in record['industries_cn'].split('\n')]
        record['solicitor_en'] = [item for item in record['solicitor_en'].split('\n\n')]
        record['solicitor_cn'] = [item for item in record['solicitor_cn'].split('\n\n')]
        records.append(record)

        query = lambda x: x['en'] == record['area']
        areas = [area for area in filter(query,constants['areas'])]
        if not areas:
            cell = ws.cell(row=r,column=keys.index('area')+1)
            cell.font = Font(color=colors.WHITE)
            cell.fill = PatternFill(fill_type='solid',fgColor=colors.RED)
            print('Error found in %s%s' % (cell.column,cell.row))
            error = True

        for item in record['industries_cn']:
            query = lambda x: x['en'] == item
            industries = [industry for industry in filter(query,constants['industries'])]
            if not industries:
                cell = ws.cell(row=r,column=keys.index('industries_cn')+1)
                cell.font = Font(color=colors.WHITE)
                cell.fill = PatternFill(fill_type='solid',fgColor=colors.RED)
                error = True
                print('Error found in %s%s' % (cell.column,cell.row))
                break

    if error:
        wb.save('boom.xlsx')
        return

    ##########################################################################################
    # store into database
    ##########################################################################################
    for record in records:
        # area
        query = lambda x: x['en'] == record['area']
        areas = [area for area in filter(query,constants['areas'])]
        area = areas[0]['id']

        # industries
        query = lambda x: x['en'] in record['industries_cn']
        industries = [industry for industry in filter(query,constants['industries'])]
        industries = [industry['id'] for industry in industries]

        category = 1    # Equity capital market (IPO & Listing & share issuance)
        party = 1       # Issuer

        client = LawClient.objects.filter(name=record['name_en']).first()
        if not client:
            client = LawClient()
            client.name = record['name_en']
            client.name_cn = record['name_cn']
            client.area = area
            client.industries = industries
            client.note = record['description_en']
            client.note_cn = record['description_cn']
            client.switch_db('primary').save()

        firms = []
        for index in range(0,len(record['solicitor_en'])):
            firm = LawFirm.objects.filter(name=record['solicitor_en'][index]).first()
            if not firm:
                firm = LawFirm()
                firm.name = record['solicitor_en'][index]
                firm.name_cn = record['solicitor_cn'][index]
                firm.categories = [category]
                firm.switch_db('primary').save()

            firms.append(firm)

        title = 'Global offering of ' + record['name_en']
        title_cn = record['name_cn'] + '全球发售'

        deal = LawDeal.objects.filter(title=title).first()
        if deal:
            continue

        deal = LawDeal()
        deal.title = title
        deal.title_cn = title_cn
        deal.categories = [category]
        deal.date = datetime.datetime.strptime(record['date'],'%Y-%m-%d')
        deal.value_txt = record['value']
        deal.raw = '\n'.join([
            'code : %s' % record['code'],
            'website : %s' % record['website'],
            'url en : %s' % record['url_en'],
            'prospectus en : %s' % record['prospectus_en'],
            'solicitor en : %s' % record['solicitor_en_url'],
            'url cn : %s' % record['url_cn'],
            'prospectus cn : %s' % record['prospectus_cn'],
            'solicitor cn : %s' % record['solicitor_cn_url'],
        ])

        deal.clients.append(EmbeddedDealClient(
            client=client,
            major=1,
            party=party,
            industries=industries,
            areas=[area],
        ))

        for firm in firms:
            deal.firms.append(EmbeddedDealFirm(
                firm=firm,
                party=party,
                areas=[area],
            ))

        deal.switch_db('primary').save()
        auto_maintain(deal)
"""



def batch_cfi_input():
    ##########################################################################################
    # load constants
    ##########################################################################################
    constants = { 'categories': [], 'industries': [], 'areas': [], 'parties': [] }
    constants['categories'] = json.loads(REDIS['app'].get('law:config:categories'))
    constants['industries'] = json.loads(REDIS['app'].get('law:config:industries'))
    constants['areas'] = json.loads(REDIS['app'].get('law:config:areas'))
    constants['parties'] = json.loads(REDIS['app'].get('law:config:parties'))

    ##########################################################################################
    # validate excel records
    ##########################################################################################
    wb = load_workbook(filename='cfi.xlsx')
    ws = wb.active
    keys = [
        'code','date','name','name_en','name_cn','phone','email','address','website',
        'value','industry','description','prospectus','firm','firm_url',
    ]

    error = False
    records = []
    for r in range(2,ws.max_row+1):
        values = []
        for c in range(1,ws.max_column+1):
            cell = ws.cell(row=r,column=c)
            values.append(cell.value)

        record = dict(zip(keys,values))
        records.append(record)

        query = lambda x: x['en'] == record['industry']
        industries = [industry for industry in filter(query,constants['industries'])]
        if not industries:
            cell = ws.cell(row=r,column=keys.index('industry')+1)
            cell.font = Font(color=colors.WHITE)
            cell.fill = PatternFill(fill_type='solid',fgColor=colors.RED)
            error = True
            print('Error found in %s%s' % (cell.column,cell.row))
            break

    if error:
        wb.save('cfi.xlsx')
        return

    ##########################################################################################
    # store into database
    ##########################################################################################
    for record in records:
        # industry
        query = lambda x: x['en'] == record['industry']
        industries = [industry for industry in filter(query,constants['industries'])]
        industry = industries[0]['id']

        area = 1        # PRC
        category = 1    # Equity capital market (IPO & Listing & share issuance)
        party = 1       # Issuer

        client = LawClient.objects.filter(name=record['name_en']).first()
        if not client:
            client = LawClient()
            client.name = record['name_en']
            client.name_cn = record['name_cn']
            client.area = area
            client.industries = [industry]
            client.note_cn = record['description']
            #client.switch_db('primary').save()

        firm = LawFirm.objects.filter(name=record['firm']).first()
        if not firm:
            firm = LawFirm()
            firm.name = record['firm']
            firm.categories = [category]
            #firm.switch_db('primary').save()

        title = 'IPO of ' + record['name_en']
        title_cn = record['name_cn'] + '首次上市'

        deal = LawDeal.objects.filter(title=title).first()
        if deal:
            continue

        deal = LawDeal()
        deal.title = title
        deal.title_cn = title_cn
        deal.categories = [category]
        deal.date = datetime.datetime.strptime(record['date'],'%Y-%m-%d')
        deal.value_txt = record['value']
        deal.raw = '\n'.join([
            'code : %s' % record['code'],
            'website : %s' % record['website'],
            'phone : %s' % record['phone'],
            'email : %s' % record['email'],
            'address : %s' % record['address'],
            'url : %s' % record['firm_url'],
            record['prospectus'],
        ])

        deal.clients.append(EmbeddedDealClient(
            client=client,
            major=1,
            party=party,
            industries=industries,
            areas=[area],
        ))

        deal.firms.append(EmbeddedDealFirm(
            firm=firm,
            party=party,
            areas=[area],
        ))

        #deal.switch_db('primary').save()
        #auto_maintain(deal)



class WebData(BaseCommand):
    def handle(self, *args, **options):
        #crawl_boom()
        #crawl_cfi()
        #crawl_cbonds()
        #export_excel()
        #batch_boom_input()
        batch_cfi_input()
        pass



